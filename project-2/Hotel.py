from tkinter import *
from tkcalendar import DateEntry
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib

#sebelum digunakan, download dulu tkcalendar, openpyxl, dan pathlib
#Buatan Rafli, Faqih, Husein, Wildan

root=Tk()
root.title("Data Pemesan Kamar Hotel") #judul
root.geometry("700x550+300+200") #untuk ukuran tampilan window aplikasi di windows
root.resizable(False,False)
root.configure(bg="#326273") #warna biru tua

#untuk cek file excel ada atau tidak
file=pathlib.Path('Backened_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="Phone Number"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"
    sheet['F1']="Check In (TTTT/BB/TT)"
    sheet['G1']="Check Out (TTTT/BB/TT)"

    file.save('Backened_data.xlsx')

def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    checkin=tanggalinEntry.get_date()
    checkout=tanggaloutEntry.get_date()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)

    if name == "" or contact == "" or age == "" or checkin is None or checkout is None or gender_combobox is "Set" or address is None:
        messagebox.showerror("Error", "Harap isi form dengan benar!")
    else: file=openpyxl.load_workbook('Backened_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)
    sheet.cell(column=6,row=sheet.max_row,value=checkin)
    sheet.cell(column=7,row=sheet.max_row,value=checkout)

    file.save(r'Backened_data.xlsx')

    messagebox.showinfo('info','detail added!')

    #akan mengulang form setelah disubmit
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    gender_combobox.set('Set')
    addressEntry.delete(1.0,END)
    tanggalinEntry.delete(0,END)
    tanggaloutEntry.delete(0,END)

    

def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    gender_combobox.set('Set')
    addressEntry.delete(1.0,END)
    tanggalinEntry.delete(0,END)
    tanggaloutEntry.delete(0,END)


#icon, menggunakan ikon di dalam folder yang sama dengan file .py ini
icon_image=PhotoImage(file="png-transparent-note-with-pen-logo-computer-icons-form-register-button-miscellaneous-angle-text.png")
root.iconphoto(False,icon_image)

#heading
Label(root,text="Masukan Data Pemesan Kamar",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#Label
Label(root,text='Name',font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text='Contact No',font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text='Age',font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text='Gender',font=23,bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text='Check In',font=23,bg="#326273",fg="#fff").place(x=50,y=250)
Label(root,text='Check Out',font=23,bg="#326273",fg="#fff").place(x=370,y=250)
Label(root,text='Address',font=23,bg="#326273",fg="#fff").place(x=50,y=300)

#Entry
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root,textvariable=nameValue,width=45,bd=2,font=20)
contactEntry = Entry(root,textvariable=contactValue,width=45,bd=2,font=20)
ageEntry = Entry(root,textvariable=AgeValue,width=15,bd=2,font=20)

#Gender
gender_combobox = Combobox(root,values=['Male', 'Female'],font='arial 14',state='r',width=14)
gender_combobox.place(x=440,y=200)
gender_combobox.set('Set')

addressEntry = Text(root,width=50,height=4,bd=2)

nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addressEntry.place(x=200,y=300)

#Tanggal
# DateEntry untuk Check In (format tanggal BB/HH/TTTT)
tanggalinEntry = DateEntry(root, width=15, background="#326273", foreground='white', borderwidth=2)
tanggalinEntry.place(x=200, y=250)

# DateEntry untuk Check Out (format tanggal BB/HH/TTTT)
tanggaloutEntry = DateEntry(root, width=15, background="#326273", foreground='white', borderwidth=2)
tanggaloutEntry.place(x=520, y=250)



#tombol
Button(root,text="Submit",bg="#326273",fg="white",width=15,height=2,command=submit).place(x=200,y=400)
Button(root,text="Clear",bg="#326273",fg="white",width=15,height=2,command=clear).place(x=340,y=400)
Button(root,text="Exit",bg="#326273",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=400)

root.mainloop()